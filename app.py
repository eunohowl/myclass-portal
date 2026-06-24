"""
MahasiswaCRUD - Python Flask Backend
OOP Blueprint: Encapsulation, Inheritance, Polymorphism
"""

import json
import os
import re
import io
from flask import Flask, request, jsonify, render_template, session, send_file
from functools import wraps

import requests

# .env opsional (kalau python-dotenv terinstall & ada file .env) — kalau
# tidak ada, baris ini cukup diabaikan dan app tetap jalan normal.
try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass

app = Flask(__name__)
app.secret_key = 'mahasiswa_secret_key_2026'

DATA_FILE = os.environ.get('DATA_PATH', 'mahasiswa.json')

# ============================================================
# Konfigurasi Brevo (untuk fitur kirim email ke mahasiswa)
# Set via environment variable di Railway:
#   BREVO_API_KEY        -> API key dari dashboard Brevo
#   BREVO_SENDER_EMAIL   -> Email pengirim yang sudah diverifikasi di Brevo
#   BREVO_SENDER_NAME    -> (opsional) nama pengirim yang ditampilkan
# ============================================================
BREVO_API_KEY      = os.environ.get('BREVO_API_KEY', '')
BREVO_SENDER_EMAIL = os.environ.get('BREVO_SENDER_EMAIL', '')
BREVO_SENDER_NAME  = os.environ.get('BREVO_SENDER_NAME', 'MYCLASS — Sistem Manajemen Mahasiswa')

# ============================================================
# OOP: Class Mahasiswa (Encapsulation)
# ============================================================
class Mahasiswa:
    def __init__(self, nim: str, nama: str, prodi: str, email: str = '', ipk: float = 0.0):
        self.__nim   = nim
        self.__nama  = nama
        self.__prodi = prodi
        self.__email = email
        self.__ipk   = float(ipk)

    # --- Getters ---
    @property
    def nim(self):   return self.__nim
    @property
    def nama(self):  return self.__nama
    @property
    def prodi(self): return self.__prodi
    @property
    def email(self): return self.__email
    @property
    def ipk(self):   return self.__ipk

    # Validasi NIM: angka, minimal 5 digit
    @staticmethod
    def validate_nim(nim: str) -> bool:
        return bool(re.fullmatch(r'[0-9]{5,}', nim))

    @staticmethod
    def validate_email(email: str) -> bool:
        return bool(re.fullmatch(r'[^@]+@[^@]+\.[^@]+', email))

    @staticmethod
    def validate_ipk(ipk) -> bool:
        try:
            val = float(ipk)
            return 0.0 <= val <= 4.0
        except (ValueError, TypeError):
            return False

    # Polimorfisme: display_info
    def display_info(self) -> str:
        cls = self.__prodi.replace(' ', '')
        return (f'Instance of Class {cls} -> '
                f'"Halo saya {self.__nama}, NIM {self.__nim} '
                f'dari Prodi {self.__prodi}, IPK {self.__ipk:.2f}."')

    def to_dict(self) -> dict:
        return {
            'nim':   self.__nim,
            'nama':  self.__nama,
            'prodi': self.__prodi,
            'email': self.__email,
            'ipk':   round(self.__ipk, 2),
        }


# ============================================================
# OOP: MahasiswaManager (File I/O + Sorting Algorithms)
# ============================================================
class MahasiswaManager:
    def __init__(self, filepath: str):
        self.filepath = filepath
        self.data: list[Mahasiswa] = []
        self.load_from_file()

    # --- File I/O ---
    def load_from_file(self):
        if os.path.exists(self.filepath):
            with open(self.filepath, 'r', encoding='utf-8') as f:
                raw = json.load(f)
            self.data = [
                Mahasiswa(r['nim'], r['nama'], r['prodi'],
                          r.get('email', ''), r.get('ipk', 0.0))
                for r in raw
            ]
        else:
            seeds = [
                Mahasiswa('202601001', 'Budi Utomo',          'Sistem Informasi',    'budi.utomo@student.ac.id',       3.75),
                Mahasiswa('202601002', 'Sutan Hasibuan',      'Teknik Informatika',  'sutan.hasibuan@student.ac.id',   3.90),
                Mahasiswa('202601003', 'Citra Lestari',       'Sains Data',          'citra.lestari@student.ac.id',    3.82),
                Mahasiswa('202601004', 'Dimas Prasetyo',      'Teknik Informatika',  'dimas.prasetyo@student.ac.id',   3.55),
                Mahasiswa('202601005', 'Erina Kusuma',        'Sistem Informasi',    'erina.kusuma@student.ac.id',     3.68),
                Mahasiswa('202601006', 'Fajar Ramadhan',      'Sains Data',          'fajar.ramadhan@student.ac.id',   3.40),
                Mahasiswa('202601007', 'Gita Maharani',       'Manajemen Informatika','gita.maharani@student.ac.id',   3.77),
                Mahasiswa('202601008', 'Hendra Wijaya',       'Teknik Informatika',  'hendra.wijaya@student.ac.id',    3.22),
                Mahasiswa('202601009', 'Indah Permatasari',   'Sistem Informasi',    'indah.permatasari@student.ac.id',3.88),
                Mahasiswa('202601010', 'Joko Santoso',        'Sains Data',          'joko.santoso@student.ac.id',     3.15),
                Mahasiswa('202601011', 'Kartika Dewi',        'Teknik Informatika',  'kartika.dewi@student.ac.id',     3.93),
                Mahasiswa('202601012', 'Lukman Hakim',        'Manajemen Informatika','lukman.hakim@student.ac.id',    3.47),
                Mahasiswa('202601013', 'Maya Putri',          'Sistem Informasi',    'maya.putri@student.ac.id',       3.61),
                Mahasiswa('202601014', 'Naufal Rizky',        'Teknik Informatika',  'naufal.rizky@student.ac.id',     3.84),
                Mahasiswa('202601015', 'Olivia Santika',      'Sains Data',          'olivia.santika@student.ac.id',   3.29),
                Mahasiswa('202601016', 'Panji Nugroho',       'Teknik Informatika',  'panji.nugroho@student.ac.id',    3.72),
                Mahasiswa('202601017', 'Qisthi Aulia',        'Manajemen Informatika','qisthi.aulia@student.ac.id',    3.56),
                Mahasiswa('202601018', 'Rizky Pratama',       'Sistem Informasi',    'rizky.pratama@student.ac.id',    3.44),
                Mahasiswa('202601019', 'Sari Indah Wulandari','Sains Data',          'sari.wulandari@student.ac.id',   3.91),
                Mahasiswa('202601020', 'Taufiq Hidayat',      'Teknik Informatika',  'taufiq.hidayat@student.ac.id',   3.67),
            ]
            self.data = seeds
            self.save_to_file()

    def save_to_file(self):
        with open(self.filepath, 'w', encoding='utf-8') as f:
            json.dump([m.to_dict() for m in self.data], f, indent=2, ensure_ascii=False)

    # --- CRUD ---
    def get_all(self) -> list[dict]:
        return [m.to_dict() for m in self.data]

    def find_by_nim(self, nim: str) -> Mahasiswa | None:
        return next((m for m in self.data if m.nim == nim), None)

    def search_by_nama(self, keyword: str) -> list[dict]:
        kw = keyword.lower()
        return [m.to_dict() for m in self.data if kw in m.nama.lower()]

    def add(self, nim: str, nama: str, prodi: str, email: str, ipk: float) -> dict:
        if not Mahasiswa.validate_nim(nim):
            raise ValueError('Format NIM tidak valid! Harus angka minimal 5 digit.')
        if self.find_by_nim(nim):
            raise ValueError(f'NIM {nim} sudah terdaftar.')
        if email and not Mahasiswa.validate_email(email):
            raise ValueError('Format email tidak valid.')
        if not Mahasiswa.validate_ipk(ipk):
            raise ValueError('IPK harus antara 0.00 dan 4.00.')
        mhs = Mahasiswa(nim, nama, prodi, email, ipk)
        self.data.append(mhs)
        self.save_to_file()
        return mhs.to_dict()

    def update(self, nim: str, nama: str, prodi: str, email: str, ipk: float) -> dict:
        mhs = self.find_by_nim(nim)
        if not mhs:
            raise ValueError(f'Mahasiswa NIM {nim} tidak ditemukan.')
        if email and not Mahasiswa.validate_email(email):
            raise ValueError('Format email tidak valid.')
        if not Mahasiswa.validate_ipk(ipk):
            raise ValueError('IPK harus antara 0.00 dan 4.00.')
        self.data = [m for m in self.data if m.nim != nim]
        updated = Mahasiswa(nim, nama, prodi, email, ipk)
        self.data.append(updated)
        self.save_to_file()
        return updated.to_dict()

    def delete(self, nim: str) -> bool:
        before = len(self.data)
        self.data = [m for m in self.data if m.nim != nim]
        if len(self.data) == before:
            raise ValueError(f'Mahasiswa NIM {nim} tidak ditemukan.')
        self.save_to_file()
        return True

    # --- Sorting Algorithms ---
    def merge_sort_by_nim(self):
        """Merge Sort — O(n log n) — by NIM"""
        self.data = self._merge_sort(self.data, key=lambda m: m.nim)
        self.save_to_file()

    def shell_sort_by_nama(self):
        """Shell Sort — by Nama"""
        arr = self.data[:]
        n = len(arr)
        gap = n // 2
        while gap > 0:
            for i in range(gap, n):
                temp = arr[i]
                j = i
                while j >= gap and arr[j - gap].nama.lower() > temp.nama.lower():
                    arr[j] = arr[j - gap]
                    j -= gap
                arr[j] = temp
            gap //= 2
        self.data = arr
        self.save_to_file()

    def _merge_sort(self, arr, key):
        if len(arr) <= 1:
            return arr
        mid = len(arr) // 2
        left  = self._merge_sort(arr[:mid], key)
        right = self._merge_sort(arr[mid:], key)
        return self._merge(left, right, key)

    def _merge(self, left, right, key):
        result = []
        i = j = 0
        while i < len(left) and j < len(right):
            if key(left[i]) <= key(right[j]):
                result.append(left[i]); i += 1
            else:
                result.append(right[j]); j += 1
        result.extend(left[i:])
        result.extend(right[j:])
        return result

    # --- Export ---
    def export_to_excel_bytes(self) -> bytes:
        """Export data to Excel (.xlsx) using openpyxl"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise ImportError('openpyxl not installed. Run: pip install openpyxl')

        wb = Workbook()
        ws = wb.active
        ws.title = 'Data Mahasiswa'

        # Header style
        header_font  = Font(bold=True, color='FFFFFF', size=11)
        header_fill  = PatternFill(start_color='1a472a', end_color='1a472a', fill_type='solid')
        center_align = Alignment(horizontal='center', vertical='center')
        thin_border  = Border(
            left   = Side(style='thin', color='D0D0D0'),
            right  = Side(style='thin', color='D0D0D0'),
            top    = Side(style='thin', color='D0D0D0'),
            bottom = Side(style='thin', color='D0D0D0'),
        )

        headers = ['No', 'NIM', 'Nama', 'Program Studi', 'Email', 'IPK']
        col_widths = [5, 14, 28, 26, 34, 8]

        for col_idx, (h, w) in enumerate(zip(headers, col_widths), start=1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = center_align
            cell.border    = thin_border
            ws.column_dimensions[get_column_letter(col_idx)].width = w

        ws.row_dimensions[1].height = 22

        # Data rows
        alt_fill = PatternFill(start_color='F0FFF4', end_color='F0FFF4', fill_type='solid')
        for row_idx, m in enumerate(self.data, start=2):
            row_data = [row_idx - 1, m.nim, m.nama, m.prodi, m.email, m.ipk]
            fill = alt_fill if row_idx % 2 == 0 else None
            for col_idx, val in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.border    = thin_border
                cell.alignment = center_align if col_idx in (1, 2, 6) else Alignment(vertical='center')
                if fill:
                    cell.fill = fill

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.read()

    # --- Import ---
    def import_from_excel_bytes(self, file_bytes: bytes, mode: str = 'skip') -> dict:
        """Import data mahasiswa dari file Excel (.xlsx) menggunakan openpyxl.

        Header kolom yang dikenali (urutan bebas, tidak peka huruf besar/kecil):
            NIM, Nama, Program Studi (atau Prodi), Email (opsional), IPK (opsional)

        mode:
            'skip'   -> baris dengan NIM yang sudah terdaftar akan dilewati
            'update' -> baris dengan NIM yang sudah terdaftar akan menimpa data lama

        Return dict ringkasan hasil import (added, updated, skipped, errors, total_rows).
        """
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise ImportError('openpyxl not installed. Run: pip install openpyxl')

        try:
            wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
        except Exception as e:
            raise ValueError(f'File Excel tidak valid atau rusak: {e}')

        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            raise ValueError('File Excel kosong, tidak ada data untuk diimpor.')

        header_row = [str(c).strip().lower() if c is not None else '' for c in rows[0]]

        def find_col(*names):
            for i, h in enumerate(header_row):
                if h in names:
                    return i
            return None

        col_nim   = find_col('nim')
        col_nama  = find_col('nama')
        col_prodi = find_col('program studi', 'prodi')
        col_email = find_col('email')
        col_ipk   = find_col('ipk')

        if col_nim is None or col_nama is None or col_prodi is None:
            raise ValueError(
                'Header kolom tidak dikenali. Pastikan baris pertama berisi '
                'header: NIM, Nama, Program Studi (Email & IPK opsional).'
            )

        def cell(row, col):
            if col is None or col >= len(row):
                return None
            return row[col]

        added, updated, skipped = 0, 0, 0
        errors = []
        total_rows = 0

        for excel_row_no, row in enumerate(rows[1:], start=2):
            if row is None or all(c is None for c in row):
                continue  # baris kosong, lewati tanpa dihitung error
            total_rows += 1

            raw_nim = cell(row, col_nim)
            nim = str(raw_nim).strip() if raw_nim is not None else ''
            if nim.endswith('.0'):  # antisipasi NIM yang terbaca sebagai float oleh Excel
                nim = nim[:-2]

            raw_nama = cell(row, col_nama)
            nama = str(raw_nama).strip() if raw_nama is not None else ''

            raw_prodi = cell(row, col_prodi)
            prodi = str(raw_prodi).strip() if raw_prodi is not None else ''

            raw_email = cell(row, col_email)
            email = str(raw_email).strip() if raw_email is not None else ''

            raw_ipk = cell(row, col_ipk)
            ipk = raw_ipk if raw_ipk is not None else 0.0

            if not nim or not nama or not prodi:
                errors.append({'row': excel_row_no, 'reason': 'NIM, Nama, atau Program Studi kosong.'})
                continue
            if not Mahasiswa.validate_nim(nim):
                errors.append({'row': excel_row_no, 'nim': nim, 'reason': 'Format NIM tidak valid (harus angka minimal 5 digit).'})
                continue
            if email and not Mahasiswa.validate_email(email):
                errors.append({'row': excel_row_no, 'nim': nim, 'reason': 'Format email tidak valid.'})
                continue
            if not Mahasiswa.validate_ipk(ipk):
                errors.append({'row': excel_row_no, 'nim': nim, 'reason': 'IPK harus antara 0.00 dan 4.00.'})
                continue

            existing = self.find_by_nim(nim)
            if existing:
                if mode == 'update':
                    self.data = [m for m in self.data if m.nim != nim]
                    self.data.append(Mahasiswa(nim, nama, prodi, email, ipk))
                    updated += 1
                else:
                    skipped += 1
                continue

            self.data.append(Mahasiswa(nim, nama, prodi, email, ipk))
            added += 1

        if added or updated:
            self.save_to_file()

        return {
            'added': added,
            'updated': updated,
            'skipped': skipped,
            'errors': errors,
            'total_rows': total_rows,
        }


# ============================================================
# EmailService — Kirim notifikasi/pesan ke mahasiswa via Brevo API
# ============================================================
class EmailService:
    """Abstraction layer untuk pengiriman email menggunakan Brevo HTTP API."""

    BREVO_URL = 'https://api.brevo.com/v3/smtp/email'

    @staticmethod
    def is_configured() -> bool:
        return bool(BREVO_API_KEY and BREVO_SENDER_EMAIL)

    @staticmethod
    def build_html(nama: str, message: str) -> str:
        safe_message = message.replace('\n', '<br>')
        return f"""
        <div style="font-family:Arial,Helvetica,sans-serif;max-width:480px;margin:0 auto;
                    border:1px solid #e5e7eb;border-radius:14px;overflow:hidden">
          <div style="background:#1a472a;color:#ffffff;padding:18px 22px;
                      font-weight:700;font-size:15px">🎓 MYCLASS — Sistem Manajemen Mahasiswa</div>
          <div style="padding:22px;color:#1f2937;line-height:1.7;font-size:14px">
            {safe_message}
          </div>
          <div style="padding:14px 22px;background:#f3f4f6;color:#9ca3af;font-size:11px">
            Email ini dikirim otomatis oleh sistem MyClass, mohon tidak membalas pesan ini.
          </div>
        </div>"""

    @classmethod
    def send(cls, to_email: str, subject: str, message: str, nama: str = '') -> None:
        if not cls.is_configured():
            raise RuntimeError(
                'Pengiriman email belum dikonfigurasi. Set environment variable '
                'BREVO_API_KEY dan BREVO_SENDER_EMAIL di Railway.'
            )
        if not Mahasiswa.validate_email(to_email):
            raise ValueError('Alamat email mahasiswa tidak valid.')

        html_body = cls.build_html(nama or to_email, message)

        payload = {
            'sender': {'name': BREVO_SENDER_NAME, 'email': BREVO_SENDER_EMAIL},
            'to': [{'email': to_email, 'name': nama or to_email}],
            'subject': subject,
            'htmlContent': html_body,
            'textContent': message,
        }
        headers = {
            'accept': 'application/json',
            'api-key': BREVO_API_KEY,
            'content-type': 'application/json',
        }

        try:
            resp = requests.post(cls.BREVO_URL, json=payload, headers=headers, timeout=10)
            if resp.status_code not in (200, 201):
                raise RuntimeError(
                    f'Brevo error {resp.status_code}: {resp.text}'
                )
        except requests.exceptions.Timeout:
            raise RuntimeError('Koneksi ke Brevo timeout. Coba lagi beberapa saat.')
        except requests.exceptions.RequestException as e:
            raise RuntimeError(f'Gagal menghubungi Brevo API: {e}')


# ============================================================
# Singleton Manager Instance
# ============================================================
manager = MahasiswaManager(DATA_FILE)

# ============================================================
# Auth Decorator
# ============================================================
USERS = {'admin': 'adminsalah123'}

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get('logged_in'):
            return jsonify({'error': 'Unauthorized'}), 401
        return f(*args, **kwargs)
    return decorated

# ============================================================
# Routes — Auth
# ============================================================
@app.route('/')
def index():
    return render_template('index.html')

@app.route('/api/login', methods=['POST'])
def api_login():
    body = request.get_json()
    username = body.get('username', '')
    password = body.get('password', '')
    if USERS.get(username) == password:
        session['logged_in'] = True
        session['username']  = username
        return jsonify({'success': True, 'username': username})
    return jsonify({'success': False, 'error': 'Username atau password salah.'}), 401

@app.route('/api/logout', methods=['POST'])
def api_logout():
    session.clear()
    return jsonify({'success': True})

# ============================================================
# Routes — CRUD
# ============================================================
@app.route('/api/mahasiswa', methods=['GET'])
@login_required
def get_all():
    return jsonify(manager.get_all())

@app.route('/api/mahasiswa/search', methods=['GET'])
@login_required
def search():
    keyword = request.args.get('nama', '')
    results = manager.search_by_nama(keyword)
    return jsonify(results)

@app.route('/api/mahasiswa/<nim>', methods=['GET'])
@login_required
def get_one(nim):
    mhs = manager.find_by_nim(nim)
    if not mhs:
        return jsonify({'error': f'NIM {nim} tidak ditemukan.'}), 404
    data = mhs.to_dict()
    data['display_info'] = mhs.display_info()
    return jsonify(data)

@app.route('/api/mahasiswa', methods=['POST'])
@login_required
def create():
    body = request.get_json()
    try:
        data = manager.add(
            body['nim'], body['nama'], body['prodi'],
            body.get('email', ''), body.get('ipk', 0.0)
        )
        return jsonify(data), 201
    except ValueError as e:
        return jsonify({'error': str(e)}), 400

@app.route('/api/mahasiswa/<nim>', methods=['PUT'])
@login_required
def update(nim):
    body = request.get_json()
    try:
        data = manager.update(
            nim, body['nama'], body['prodi'],
            body.get('email', ''), body.get('ipk', 0.0)
        )
        return jsonify(data)
    except ValueError as e:
        return jsonify({'error': str(e)}), 400

@app.route('/api/mahasiswa/<nim>', methods=['DELETE'])
@login_required
def delete(nim):
    try:
        manager.delete(nim)
        return jsonify({'success': True})
    except ValueError as e:
        return jsonify({'error': str(e)}), 404

@app.route('/api/mahasiswa/sort/<method>', methods=['POST'])
@login_required
def sort_data(method):
    if method == 'merge':
        manager.merge_sort_by_nim()
        return jsonify({'message': 'Diurutkan berdasarkan NIM (Merge Sort)', 'data': manager.get_all()})
    elif method == 'shell':
        manager.shell_sort_by_nama()
        return jsonify({'message': 'Diurutkan berdasarkan Nama (Shell Sort)', 'data': manager.get_all()})
    return jsonify({'error': 'Metode tidak dikenal.'}), 400

@app.route('/api/export/excel', methods=['GET'])
@login_required
def export_excel():
    try:
        excel_bytes = manager.export_to_excel_bytes()
        return send_file(
            io.BytesIO(excel_bytes),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='data_mahasiswa.xlsx'
        )
    except ImportError as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/import/excel', methods=['POST'])
@login_required
def import_excel():
    if 'file' not in request.files:
        return jsonify({'error': 'Tidak ada file yang diunggah.'}), 400
    file = request.files['file']
    if not file or file.filename == '':
        return jsonify({'error': 'Tidak ada file yang dipilih.'}), 400
    if not file.filename.lower().endswith(('.xlsx', '.xlsm')):
        return jsonify({'error': 'Format file harus .xlsx atau .xlsm.'}), 400

    mode = request.form.get('mode', 'skip')
    if mode not in ('skip', 'update'):
        mode = 'skip'

    try:
        file_bytes = file.read()
        result = manager.import_from_excel_bytes(file_bytes, mode)
        return jsonify({'success': True, **result})
    except ImportError as e:
        return jsonify({'error': str(e)}), 500
    except ValueError as e:
        return jsonify({'error': str(e)}), 400
    except Exception as e:
        return jsonify({'error': f'Gagal mengimpor file: {e}'}), 500


@app.route('/api/mahasiswa/<nim>/email', methods=['POST'])
@login_required
def send_email(nim):
    mhs = manager.find_by_nim(nim)
    if not mhs:
        return jsonify({'error': f'Mahasiswa NIM {nim} tidak ditemukan.'}), 404
    if not mhs.email:
        return jsonify({'error': f'Mahasiswa {mhs.nama} tidak memiliki alamat email.'}), 400

    body = request.get_json() or {}
    subject = (body.get('subject') or '').strip() or f'Informasi Akademik — {mhs.nama}'
    message = (body.get('message') or '').strip()
    if not message:
        message = (
            f'Halo {mhs.nama},\n\n'
            f'Berikut informasi akademik Anda saat ini:\n'
            f'NIM: {mhs.nim}\nProgram Studi: {mhs.prodi}\nIPK: {mhs.ipk:.2f}\n\n'
            f'Terima kasih.'
        )

    try:
        EmailService.send(mhs.email, subject, message, nama=mhs.nama)
        return jsonify({'success': True, 'message': f'Email berhasil dikirim ke {mhs.email}'})
    except (RuntimeError, ValueError) as e:
        return jsonify({'error': str(e)}), 400
    except Exception as e:
        return jsonify({'error': f'Terjadi kesalahan tidak terduga: {e}'}), 500


if __name__ == '__main__':
    print("=" * 50)
    print("  MahasiswaCRUD - Flask Server")
    print("  Buka: http://127.0.0.1:5000")
    print("=" * 50)
    app.run(debug=True)
